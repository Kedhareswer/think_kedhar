"""Master sheet exporter.

Produces an .xlsx file mirroring Master sheet.xlsx — same nine sheets, same
column headers, same order — populated from the medbrain corpus.

The exporter is intentionally additive and idempotent: each run rewrites the
file from scratch. Empty sheets keep their headers so downstream consumers
see a stable schema even before data lands for that domain.

Column headers are reproduced verbatim from the reference file (including
trailing spaces, double spaces, and capitalisation) so a diff against the
original is meaningful for human reviewers.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

from sqlalchemy import or_, select

from .. import config
from ..db import session_scope
from ..enums import ClaimStatus
from ..models import Claim, Source

# Reference-style prompt paths — one per sheet that has a row builder.
_PROMPTS_DIR = Path(__file__).resolve().parent.parent.parent / "prompts"
_LLM_PROMPT_PATH = _PROMPTS_DIR / "master_sheet_condition.md"
_LLM_MED_PROMPT_PATH = _PROMPTS_DIR / "master_sheet_medication.md"

# Slug substrings that identify a concept as a medication (used by the
# Medications row builder to pick out drug concepts).
_MEDICATION_SLUG_HINTS: tuple[str, ...] = (
    "artemether", "lumefantrine", "chloroquine", "primaquine", "mefloquine",
    "quinine", "doxycycline", "tafenoquine", "piperaquine", "amodiaquine",
    "artesunate", "artefenomel", "sulfadoxine", "pyrimethamine",
    "artemisinin", "oz277",
)


def _looks_like_medication(slug: str) -> bool:
    s = slug.lower()
    return any(h in s for h in _MEDICATION_SLUG_HINTS)

# ---------------------------------------------------------------------------
# Sheet schemas — header text matches Master sheet.xlsx character-for-character
# ---------------------------------------------------------------------------

CONDITIONS_COLS: list[str] = [
    "Completed", "Reviewed", "System", "Condition", "Overview", "Symptoms",
    "When to See a Doctor", "Causes", "Risk Factors", "Complications",
    "Prevention", "Diagnosis", "Treatment", "Outlook/Prognosis",
    "The Gist", "Tags", "Sources",
]

MEDICATIONS_COLS: list[str] = [
    "THERAPEUTIC CATEGORY", "DRUG GROUP ", "DRUG CLASS", "DRUG/GENERIC NAME\n\n",
    "NEO  URL", "OVERVIEW", "INDICATION (Why is this medication prescribed?)",
    "MECHANISM OF ACTION (How does it work?)",
    "DOSAGE AND ADMINISTRATION (How should this medicine be used?)",
    "CONTRAINDICATIONS (When not to use it)", "WARNINGS AND PRECAUTION\n ",
    "SIDE EFFECTS ",
    "DRUG INTERACTIONS (How this drug affects other drugs or is affected by other drugs)",
    "CONTROLLED SUBSTANCE", "STORAGE AND STABILITY ", "TAGS", "SOURCES",
    "THE GIST", "BRAND NAMES in Nigeria", "NAFDAC number ",
    "Price range in naira", "BRAND NAMES in UK", "BRAND NAMES in US",
    "BRAND NAMES in South Africa", "BRAND NAMES in Canada",
]

LAB_TESTS_COLS: list[str] = [
    "Body Systems", "Category", "Completed", "Reviewed", "Test", "Description ",
    "Purpose - What is the test used for?",
    "Indications - When do you need the test?", "Sample Required",
    "Test Preparation Needed",
    "Interpretation - What do the reults mean?",
    "Tags", "Sources", "This Gist ", "Pricing  range",
]

RADIOLOGY_COLS: list[str] = [
    "GROUP ", "TEST", "DESCRIPTION ", "PURPOSE - WHAT IS THE TEST USED FOR?",
    "INDICATIONS- WHEN DO YOU NEED THE TEST?", "PREPARATION GUIDELINES",
    "PROCEDURE DETAILS (HOW IT IS DONE)", "INTERPRETATION ",
    "RISKS AND CONSIDERATIONS", "FOLLOW-UP ACTIONS", "TAGS", "SOURCES",
    "THE GIST", "PRICE RANGE",
]

SPECIAL_DIAG_COLS: list[str] = [
    "Group", "Test", "Completed", "Reviewed", "Description ",
    "Purpose - What is the test used for?",
    "Indications - When do you need the test?", "Test Preparation Needed",
    "Procedure Details (How it is done)",
    "Interpretation - What do the reults mean?",
    "Tags", "Sources", "This Gist ", "Pricing  range",
]

HISTOPATH_COLS: list[str] = [
    "Group", "Test", "Completed", "Reviewed", "Description ",
    "Purpose - What is the test used for?",
    "Indications - When do you need the test?", "Test Preparation Needed",
    "Procedure Details (How it is done)",
    "Interpretation - What do the reults mean?",
    "Tags", "Sources", "Pricing  range", "This Gist ",
]

SURGERIES_COLS: list[str] = [
    "Completed", "Reviewed", "Top Level Category", "Surgery/Procedure",
    "Overview", "Why the Procedure is Performed",
    "Risks and Possible Complications\n",
    " Preparation Before Procedure", "How is the Procedure Performed",
    "Recovery and Aftercare", "Emergency Post-op Complications ",
    "Expected Outcomes ", "The Gist", "Tags", "Sources",
]

PROVIDERS_COLS: list[str] = [
    "Provider Type", "Specialty", "Subspecialty", "Description",
    "Services Rendered",
]

HOSPITALS_COLS: list[str] = [
    "State/Province", "Facility Name", "Verified", "Type", "Care Level", "LGA",
    "City/District", "Ward", "Neighbourhood ", "Address Line 1",
    "Address Line 2", "Latitude", "Longitude", "Facility Admin Phone Number ",
    "Opening hours (Mon- Fri)", "Opening hours (Saturdays)",
    "Opening hours (Sundays)", "Services Rendered",
]


@dataclass(frozen=True)
class SheetSpec:
    name: str
    columns: list[str]
    rows_builder: Callable[[], Iterable[dict[str, str]]] | None = None


# ---------------------------------------------------------------------------
# Markdown section extraction — best-effort mapping from concept.md to columns
# ---------------------------------------------------------------------------

# Aliases let the same concept be filed under different section headings while
# still landing in the right Master-sheet column. Both the typed-prompt
# headings (e.g. "Overview") and the legacy research-synthesis headings
# (e.g. "Synthesis", "Core claim") are accepted as Overview sources so
# pre-typed-prompt concepts still surface something useful.
CONDITION_SECTION_ALIASES: dict[str, tuple[str, ...]] = {
    "Overview":            ("overview", "summary", "definition", "core assertion",
                            "synthesis", "synthesis statement", "core claim",
                            "core claims", "what it is", "context", "bottom line"),
    "Symptoms":            ("symptoms", "clinical features", "presentation"),
    "When to See a Doctor": ("when to see a doctor", "red flags", "warning signs"),
    "Causes":              ("causes", "etiology", "aetiology", "pathogenesis",
                            "mechanism", "mechanism logic", "mechanistic logic"),
    "Risk Factors":        ("risk factors", "risk", "epidemiology"),
    "Complications":       ("complications", "sequelae", "implications",
                            "clinical relevance"),
    "Prevention":          ("prevention", "prophylaxis"),
    "Diagnosis":           ("diagnosis", "investigations", "workup"),
    "Treatment":           ("treatment", "management", "therapy", "efficacy"),
    "Outlook/Prognosis":   ("outlook", "prognosis", "outlook/prognosis", "status"),
}

# Approximate body-system mapping; expand as concepts grow.
SYSTEM_HINTS: list[tuple[str, str]] = [
    # Malaria-domain markers — all map to Infectious Disease for v1.
    ("malaria",     "Infectious Disease"),
    ("plasmodium",  "Infectious Disease"),
    ("falciparum",  "Infectious Disease"),
    ("vivax",       "Infectious Disease"),
    ("artemisinin", "Infectious Disease"),
    ("artemether",  "Infectious Disease"),
    ("lumefantrine","Infectious Disease"),
    ("chloroquine", "Infectious Disease"),
    ("kelch13",     "Infectious Disease"),
    ("pfkelch13",   "Infectious Disease"),
    ("pfcoronin",   "Infectious Disease"),
    ("anopheles",   "Infectious Disease"),
    ("act-",        "Infectious Disease"),
    ("art-",        "Infectious Disease"),
    ("acts-",       "Infectious Disease"),
    ("mefloquine",  "Infectious Disease"),
    ("primaquine",  "Infectious Disease"),
    ("quinine",     "Infectious Disease"),
    ("ipt",         "Infectious Disease"),
    ("sp-",         "Infectious Disease"),
]


def _system_for(slug: str) -> str:
    s = slug.lower()
    for needle, system in SYSTEM_HINTS:
        if needle in s:
            return system
    return ""


_HEADING_RE = re.compile(r"^(#{1,6})\s+(.*?)\s*$")


def _parse_sections(md: str) -> dict[str, str]:
    """Split markdown into {heading_text_lower: body_text} sections.

    Headings at any depth are flattened. Duplicate headings concatenate so no
    content is silently dropped — the resulting body for a key is the join of
    every section that ever appeared under it.
    """
    out: dict[str, list[str]] = {}
    current = "_intro"
    for line in md.splitlines():
        m = _HEADING_RE.match(line)
        if m:
            current = m.group(2).strip().lower()
            out.setdefault(current, [])
        else:
            out.setdefault(current, []).append(line)
    return {k: "\n".join(v).strip() for k, v in out.items()}


# Slug substrings that strongly imply the concept is NOT a stand-alone
# condition (it belongs on Medications, a gene/mutation register, etc.).
# These rows are skipped so the Conditions sheet stays clean for reviewers;
# they'll resurface on the Medications sheet once that builder lands.
_NON_CONDITION_SLUG_HINTS: tuple[str, ...] = (
    # antimalarial drugs and other drug-name roots that surface here
    "artemether", "lumefantrine", "chloroquine", "primaquine", "mefloquine",
    "quinine", "doxycycline", "tafenoquine", "piperaquine", "amodiaquine",
    "artesunate", "artefenomel", "sulfadoxine", "pyrimethamine",
    "artemisinin", "metformin", "paclitaxel", "oz277",
    # gene / protein / parasite-machinery names
    "kelch13", "pfkelch13", "pfcoronin", "pfk13", "pfmdr",
    "pfatg", "pfcrt", "pfpm",
    # mutations
    "mutation", "isoform", "c580y", "r539t", "r561h", "y493h", "y561h",
    "f476i", "g50e", "e107v", "t38i",
    # mechanism / pathway / regimen labels
    "act-",            # "ACT-overall-efficacy-maintenance" etc.
    "ipt", "rdt",
    "endocytosis", "er-stress", "hemoglobin-acquisition",
    # lab-test / category labels (belong on Lab tests / Special Diagnostics)
    "assay", "cytometric",
    # generic safety / outcomes (not stand-alone conditions)
    "adverse-events",
)


def _looks_like_condition(slug: str, sections: dict[str, str]) -> bool:
    """Decide whether ``slug`` belongs on the Conditions sheet.

    Strategy: exclude obvious non-conditions (drugs, genes, mutations,
    therapy regimens) by slug substring. Whatever remains is treated as
    condition-shaped and lands on the sheet. This is more permissive than
    requiring specific section headings, which the legacy
    research-synthesis concept notes don't carry.
    """
    s = slug.lower()
    for hint in _NON_CONDITION_SLUG_HINTS:
        if hint in s:
            return False
    return True


def _pick_section(sections: dict[str, str], aliases: tuple[str, ...]) -> str:
    for alias in aliases:
        if alias in sections:
            return sections[alias]
    # fuzzy: contains
    for key, body in sections.items():
        if any(a in key for a in aliases):
            return body
    return ""


def _humanise(slug: str) -> str:
    return slug.replace("-", " ").replace("_", " ").title()


def _read(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError):
        return ""


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------

def _conditions_rows() -> Iterable[dict[str, str]]:
    """Emit one row per concept whose .md looks like a medical condition.

    Drugs, mechanisms, gene mutations, and other non-condition concepts are
    skipped so the Conditions sheet stays clean for human reviewers. They'll
    land on Medications / Genes / etc. sheets once those builders exist.
    """
    if not config.CONCEPTS_DIR.exists():
        return []
    rows: list[dict[str, str]] = []
    for path in sorted(config.CONCEPTS_DIR.glob("*.md")):
        slug = path.stem
        md = _read(path)
        sections = _parse_sections(md)
        if not _looks_like_condition(slug, sections):
            continue
        gist_path = config.MEMORY_DIR / f"{slug}.md"
        gist = _read(gist_path).strip()
        # Fall back to the synthesis heading as a Gist-of-last-resort when
        # Brain hasn't written a per-concept memory yet.
        if not gist:
            gist = _pick_section(sections, ("synthesis", "synthesis statement", "the gist"))
        row: dict[str, str] = {c: "" for c in CONDITIONS_COLS}
        row["Completed"] = "No"
        row["Reviewed"] = "No"
        row["System"] = _system_for(slug)
        row["Condition"] = _humanise(slug)
        for col, aliases in CONDITION_SECTION_ALIASES.items():
            row[col] = _pick_section(sections, aliases)
        row["The Gist"] = gist
        rows.append(row)
    return rows


# --------------------------------------------------------------------------
# LLM-rich row builder (--llm flag) — uses Claude CLI to produce reference-
# style prose per cell. Slow (one LLM call per concept) but produces rows
# the human reviewer can actually use without further editing.
# --------------------------------------------------------------------------


def _claims_for(entity_slug: str) -> list[dict[str, object]]:
    """Pull all non-rejected claims that name the entity (subject or object)."""
    needle = entity_slug.replace("-", " ").strip()
    if not needle:
        return []
    payload: list[dict[str, object]] = []
    with session_scope() as sess:
        rows = sess.execute(
            select(Claim, Source)
            .join(Source, Claim.source_id == Source.source_id, isouter=True)
            .where(
                or_(
                    Claim.subject_text.ilike(f"%{needle}%"),
                    Claim.object_text.ilike(f"%{needle}%"),
                ),
                Claim.status != ClaimStatus.REJECTED,
            )
            .order_by(Claim.ingested_at.desc())
        ).all()
        for c, s in rows:
            payload.append({
                "claim_id": c.claim_id[:8],
                "subject": c.subject_text,
                "predicate": c.predicate.value,
                "object": c.object_text,
                "qualifiers": c.qualifiers or {},
                "certainty": c.certainty.value,
                "evidence_grade": c.evidence_grade.value,
                "source": {
                    "type": s.source_type.value if s else None,
                    "external_id": s.external_id if s else None,
                    "title": s.title if s else None,
                } if s else None,
            })
    return payload


def _llm_condition_row(slug: str, concept_md: str, gist_md: str) -> dict[str, str] | None:
    """Invoke the LLM to produce a Master-sheet-style row for one condition.

    Returns None if the LLM call fails — caller falls back to the fast
    section-extractor row so the export never breaks because of one bad cell.
    """
    try:
        from ..llm import LLMError, call_json
    except ImportError:
        return None

    claims = _claims_for(slug)
    if not concept_md and not claims:
        return None

    system = _LLM_PROMPT_PATH.read_text(encoding="utf-8")
    user = (
        f"# Entity slug\n{slug}\n\n"
        f"# Concept .md (existing)\n\n{concept_md or '(empty)'}\n\n"
        f"# Brain memory / gist (existing)\n\n{gist_md or '(empty)'}\n\n"
        f"# Claims ({len(claims)})\n```json\n"
        f"{json.dumps(claims, indent=2, default=str)}\n```\n"
    )
    try:
        obj = call_json(system, user, timeout=180.0)
    except LLMError:
        return None
    if not isinstance(obj, dict):
        return None
    # Coerce every column to a string and fill missing ones with "".
    return {col: str(obj.get(col, "") or "") for col in CONDITIONS_COLS}


def _llm_medication_row(slug: str, concept_md: str) -> dict[str, str] | None:
    """LLM Master-sheet row for one drug concept."""
    try:
        from ..llm import LLMError, call_json
    except ImportError:
        return None

    claims = _claims_for(slug)
    if not concept_md and not claims:
        return None

    system = _LLM_MED_PROMPT_PATH.read_text(encoding="utf-8")
    user = (
        f"# Drug slug\n{slug}\n\n"
        f"# Concept .md (existing)\n\n{concept_md or '(empty)'}\n\n"
        f"# Claims ({len(claims)})\n```json\n"
        f"{json.dumps(claims, indent=2, default=str)}\n```\n"
    )
    try:
        obj = call_json(system, user, timeout=180.0)
    except LLMError:
        return None
    if not isinstance(obj, dict):
        return None
    return {col: str(obj.get(col, "") or "") for col in MEDICATIONS_COLS}


def medications_rows_llm() -> Iterable[dict[str, str]]:
    """LLM-rich Medications row builder. One Claude call per drug concept."""
    if not config.CONCEPTS_DIR.exists():
        return []
    rows: list[dict[str, str]] = []
    for path in sorted(config.CONCEPTS_DIR.glob("*.md")):
        slug = path.stem
        if not _looks_like_medication(slug):
            continue
        row = _llm_medication_row(slug, _read(path))
        if row is not None:
            rows.append(row)
    return rows


def conditions_rows_llm() -> Iterable[dict[str, str]]:
    """LLM-rich variant of ``_conditions_rows`` — one Claude call per concept."""
    if not config.CONCEPTS_DIR.exists():
        return []
    rows: list[dict[str, str]] = []
    for path in sorted(config.CONCEPTS_DIR.glob("*.md")):
        slug = path.stem
        if not _looks_like_condition(slug, _parse_sections(_read(path))):
            continue
        gist = _read(config.MEMORY_DIR / f"{slug}.md")
        row = _llm_condition_row(slug, _read(path), gist)
        if row is None:
            # fall back to the fast extractor so this row isn't lost
            md = _read(path)
            sections = _parse_sections(md)
            row = {c: "" for c in CONDITIONS_COLS}
            row["Completed"] = "No"
            row["Reviewed"] = "No"
            row["System"] = _system_for(slug)
            row["Condition"] = _humanise(slug)
            for col, aliases in CONDITION_SECTION_ALIASES.items():
                row[col] = _pick_section(sections, aliases)
            row["The Gist"] = gist.strip() or _pick_section(
                sections, ("synthesis", "synthesis statement", "the gist", "summary")
            )
        rows.append(row)
    return rows


# Domains we don't yet have data for — return empty so the sheet keeps its
# headers but stays empty. Wire real builders here as new domains come online.
def _empty() -> Iterable[dict[str, str]]:
    return []


SHEETS: list[SheetSpec] = [
    SheetSpec("Conditions",                  CONDITIONS_COLS,  _conditions_rows),
    SheetSpec("Medications",                 MEDICATIONS_COLS, _empty),
    SheetSpec("Lab tests",                   LAB_TESTS_COLS,   _empty),
    SheetSpec("RadiologyImaging",            RADIOLOGY_COLS,   _empty),
    SheetSpec("Special Diagnostic Studies",  SPECIAL_DIAG_COLS, _empty),
    SheetSpec("HistopathologyCytology",      HISTOPATH_COLS,   _empty),
    SheetSpec("SurgeriesProcedures",         SURGERIES_COLS,   _empty),
    SheetSpec("Provider type & Specialty",   PROVIDERS_COLS,   _empty),
    SheetSpec("Hospital Addressbook",        HOSPITALS_COLS,   _empty),
]


# ---------------------------------------------------------------------------
# Writer
# ---------------------------------------------------------------------------

def export(output_path: Path | None = None, *, llm: bool = False) -> Path:
    """Write all nine sheets to ``output_path`` (defaults to config.MASTER_SHEET).

    Parameters:
      output_path: target xlsx; defaults to config.MASTER_SHEET.
      llm: when True, use the LLM-rich builders (currently Conditions only)
           which call Claude per concept and produce reference-style prose.
           Slower (~5–15s per concept) but the rows are reviewer-ready.

    Returns the path actually written so callers can log it.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError(
            "openpyxl is required for master sheet export. "
            "Install with: pip install openpyxl"
        ) from exc

    output_path = output_path or config.MASTER_SHEET
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # discard default sheet

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEEEEE")
    wrap = Alignment(wrap_text=True, vertical="top")

    for spec in SHEETS:
        ws = wb.create_sheet(title=spec.name[:31])  # excel hard-limits to 31 chars
        ws.append(spec.columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"

        builder = spec.rows_builder
        if llm and spec.name == "Conditions":
            builder = conditions_rows_llm
        elif llm and spec.name == "Medications":
            builder = medications_rows_llm
        rows = list(builder() if builder else [])
        for row in rows:
            ws.append([row.get(col, "") for col in spec.columns])
        # auto-ish column widths
        for i, col_name in enumerate(spec.columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(
                14, min(40, len(col_name.strip()) + 2)
            )
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(spec.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = wrap

    wb.save(output_path)
    return output_path
